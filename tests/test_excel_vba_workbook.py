from __future__ import annotations

import unittest
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "excel" / "cachexia_risk_mock_ui.v1.1.xlsm"
VBA_SOURCE = ROOT / "vba" / "CachexiaUI.bas"
VBA_SHEET_SOURCE = ROOT / "vba" / "MockUISheet.cls"
VBA_WORKBOOK_SOURCE = ROOT / "vba" / "ThisWorkbook.cls"


class ExcelVbaPrototypeTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        if not WORKBOOK.exists():
            raise AssertionError(
                "Build the workbook with "
                "python3.12 scripts/build_excel_vba_prototype.py"
            )
        cls.workbook = load_workbook(WORKBOOK, data_only=False, keep_vba=True)

    def test_workbook_is_macro_enabled_and_has_expected_sheets(self):
        self.assertIsNotNone(self.workbook.vba_archive)
        self.assertEqual(
            self.workbook.sheetnames,
            ["Mock UI", "Engine", "Assumptions", "Clinical Review"],
        )
        self.assertEqual(
            [
                sheet.title
                for sheet in self.workbook.worksheets
                if sheet.sheet_state == "visible"
            ],
            ["Mock UI", "Clinical Review"],
        )

    def test_mock_ui_has_prominent_safety_notice_and_live_outputs(self):
        sheet = self.workbook["Mock UI"]
        self.assertIn("NOT FOR CLINICAL USE", sheet["A3"].value)
        self.assertEqual(sheet["G7"].value, "3  Automatic illustrative outputs")
        self.assertEqual(sheet["G16"].value, "Current cachexia criteria status")
        self.assertEqual(
            sheet["G17"].value,
            "Current provisional pre-cachexia candidate",
        )
        self.assertNotIn("Option B", sheet["G29"].value)
        self.assertEqual(sheet["H9"].value, "=Engine!B9")
        self.assertEqual(sheet["G21"].value, "=Engine!B25")
        self.assertEqual(sheet["K21"].value, "=Engine!D25")
        self.assertEqual(
            sheet["G24"].value,
            "Target outcome is not defined pending clinical review.",
        )

    def test_inputs_have_visible_guidance_and_dynamic_subtype_validation(self):
        sheet = self.workbook["Mock UI"]
        for row in range(9, 19):
            self.assertTrue(sheet[f"D{row}"].value, f"Missing guidance at row {row}")
        self.assertIn("Synthetic stratifier", sheet["D12"].value)
        self.assertIn("descriptive and unused", sheet["D13"].value)
        self.assertIn("Optional", sheet["D15"].value)
        validations = list(sheet.data_validations.dataValidation)
        subtype = next(
            validation
            for validation in validations
            if "C13" in str(validation.sqref)
        )
        self.assertIn("LungSubtypeValues", subtype.formula1)
        self.assertIn("NonLungSubtypeValues", subtype.formula1)
        self.assertIn("LungSubtypeValues", self.workbook.defined_names)
        self.assertIn("NonLungSubtypeValues", self.workbook.defined_names)

    def test_temporal_helpers_exclude_post_prediction_weights(self):
        sheet = self.workbook["Mock UI"]
        self.assertIn("B24<=$C$9", sheet["M24"].value)
        self.assertIn("B24<Engine!$B$8", sheet["N24"].value)
        self.assertIn("EDATE(Engine!$B$8,-6)", sheet["N24"].value)

    def test_three_and_six_month_risk_formulas_are_separate(self):
        engine = self.workbook["Engine"]
        self.assertIn("Assumptions!$B$20", engine["B25"].value)
        self.assertNotIn("Assumptions!$B$25", engine["B25"].value)
        self.assertIn("Assumptions!$B$25", engine["D25"].value)
        self.assertNotIn("Assumptions!$B$20", engine["D25"].value)

    def test_sarcopenia_is_future_use_and_disabled_in_v1_formula(self):
        formula = self.workbook["Engine"]["B19"].value
        self.assertNotIn("'Mock UI'!C18", formula)
        self.assertIn('"yes","unknown"', formula)
        self.assertEqual(
            self.workbook["Mock UI"]["B18"].value,
            "Sarcopenia (future-use)",
        )

    def test_category_formulas_withhold_all_required_unknowns(self):
        engine = self.workbook["Engine"]
        formula = engine["B25"].value
        for token in (
            "'Mock UI'!C14=\"unknown\"",
            "'Mock UI'!C16=\"unknown\"",
            "'Mock UI'!C17=\"unknown\"",
            'B12=""',
            'B13=""',
        ):
            self.assertIn(token, formula)
        self.assertIn('"moderate"', formula)
        self.assertNotIn("EXP(", formula)
        self.assertEqual(self.workbook["Mock UI"]["G23"].value, "Basis / status")

    def test_form_buttons_are_present_and_wired_to_macros(self):
        with ZipFile(WORKBOOK) as archive:
            vml = "\n".join(
                archive.read(name).decode("utf-8")
                for name in archive.namelist()
                if name.endswith(".vml")
            )
            self.assertIn("CalculateRisk", vml)
            self.assertIn("ResetForm", vml)
            self.assertIn("LoadLowRiskExample", vml)
            self.assertIn("LoadHighRiskExample", vml)
            self.assertIn("OpenClinicalReview", vml)
            self.assertIn("OpenMockUI", vml)
            self.assertIn("xl/vbaProject.bin", archive.namelist())

    def test_vba_source_is_version_controlled_and_non_clinical(self):
        source = VBA_SOURCE.read_text(encoding="utf-8")
        self.assertIn("Public Sub CalculateRisk()", source)
        self.assertIn("Public Sub ResetForm()", source)
        self.assertIn("Public Sub LoadLowRiskExample()", source)
        self.assertIn("Public Sub LoadHighRiskExample()", source)
        self.assertIn("Public Sub ConfigureLungSubtypeField()", source)
        self.assertIn("Public Sub UpdateInputGuidance()", source)
        self.assertIn("must not be used for clinical decisions", source)
        self.assertIn(
            "ConfigureLungSubtypeField",
            VBA_SHEET_SOURCE.read_text(encoding="utf-8"),
        )
        self.assertIn(
            "InitializeMockUI",
            VBA_WORKBOOK_SOURCE.read_text(encoding="utf-8"),
        )


if __name__ == "__main__":
    unittest.main()
