from __future__ import annotations

import json
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "excel" / "cachexia_risk_prototype.v1.4.xlsx"


class ExcelPrototypeTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        if not WORKBOOK.exists():
            raise AssertionError(
                "Build the workbook with python3 scripts/build_excel_prototype.py"
            )
        cls.workbook = load_workbook(WORKBOOK, data_only=False)

    def test_required_interactive_sheets_exist(self):
        self.assertEqual(
            self.workbook.sheetnames,
            [
                "START HERE",
                "Calculator",
                "Results",
                "Assumptions",
                "Synthetic Cohort",
                "Data Dictionary",
                "Clinical Review",
            ],
        )
        self.assertEqual(
            [
                sheet.title
                for sheet in self.workbook.worksheets
                if sheet.sheet_state == "visible"
            ],
            ["Calculator", "Clinical Review"],
        )

    def test_safety_notice_is_prominent(self):
        for sheet in self.workbook.worksheets:
            self.assertIn("NOT FOR CLINICAL USE", str(sheet["A4"].value))

    def test_predictor_formula_enforces_prediction_date_cutoff(self):
        helper_formula = self.workbook["Calculator"]["C22"].value
        self.assertIn("A22<=$B$8", helper_formula)
        result_formula = self.workbook["Results"]["B8"].value
        self.assertIn("MAX(", result_formula)
        self.assertIn("'Calculator'!$C$22:$C$33", result_formula)

    def test_formulas_avoid_unencoded_future_functions(self):
        disallowed = ("MAXIFS(", "MINIFS(", "TEXTJOIN(")
        for sheet in self.workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        self.assertFalse(
                            any(name in cell.value for name in disallowed),
                            f"{sheet.title}!{cell.coordinate}: {cell.value}",
                        )

    def test_conservative_ooxml_avoids_fragile_features(self):
        self.assertEqual(list(self.workbook.defined_names.values()), [])
        for sheet in self.workbook.worksheets:
            self.assertEqual(list(sheet.tables.values()), [])
            self.assertEqual(len(sheet.conditional_formatting), 0)
        self.assertFalse(self.workbook.calculation.fullCalcOnLoad)
        self.assertFalse(self.workbook.calculation.forceFullCalc)

    def test_custom_unit_formats_quote_literal_text(self):
        calculator = self.workbook["Calculator"]
        self.assertEqual(calculator["H9"].number_format, '0.00 "kg"')
        self.assertEqual(calculator["H12"].number_format, '0 "days"')
        self.assertEqual(calculator["H13"].number_format, '0.00 "kg/month"')

    def test_three_and_six_month_formulas_are_separate(self):
        results = self.workbook["Results"]
        self.assertIn("'Assumptions'!$B$20", results["B25"].value)
        self.assertNotIn("'Assumptions'!$B$25", results["B25"].value)
        self.assertIn("'Assumptions'!$B$25", results["D25"].value)
        self.assertNotIn("'Assumptions'!$B$20", results["D25"].value)

    def test_unknown_is_available_in_validated_inputs(self):
        validations = list(self.workbook["Calculator"].data_validations.dataValidation)
        formulas = {validation.formula1 for validation in validations}
        self.assertTrue(any("unknown" in str(formula) for formula in formulas))
        self.assertTrue(
            any(
                "INDIRECT" in str(validation.formula1)
                and "Assumptions!$T$8:$T$10" in str(validation.formula1)
                and "Assumptions!$U$8:$U$8" in str(validation.formula1)
                for validation in validations
            )
        )

    def test_numeric_validations_have_stop_alerts(self):
        validations = list(
            self.workbook["Calculator"].data_validations.dataValidation
        )
        numeric = [
            validation
            for validation in validations
            if validation.type in {"whole", "decimal", "date"}
        ]
        self.assertTrue(numeric)
        for validation in numeric:
            self.assertTrue(validation.showErrorMessage)
            self.assertEqual(validation.errorStyle, "stop")
            self.assertTrue(validation.error)

    def test_missing_required_predictors_withhold_excel_category(self):
        results = self.workbook["Results"]
        category_formula = results["B25"].value
        explanation_formula = results["B27"].value
        for token in (
            "'Calculator'!B13=\"unknown\"",
            "'Calculator'!B15=\"unknown\"",
            "'Calculator'!B16=\"unknown\"",
            'B12=""',
            'B13=""',
        ):
            self.assertIn(token, category_formula)
        self.assertIn('"withheld"', category_formula)
        self.assertIn('"moderate"', category_formula)
        self.assertIn("Illustrative simulation category withheld", explanation_formula)

    def test_sarcopenia_branch_is_disabled_in_v1_formula(self):
        formula = self.workbook["Results"]["B19"].value
        self.assertNotIn("'Calculator'!B17", formula)
        self.assertIn("'Assumptions'!$B$12", formula)
        self.assertIn('"yes","unknown"', formula)

    def test_single_screen_exposes_inputs_and_outputs(self):
        calculator = self.workbook["Calculator"]
        self.assertEqual(
            calculator["G8"].value,
            "Baseline-derived status and illustrative categories",
        )
        self.assertEqual(calculator["H10"].value, "='Results'!B12")
        self.assertEqual(calculator["K9"].value, "='Results'!B25")
        self.assertEqual(
            calculator["K12"].value,
            "Target outcome is not defined pending clinical review.",
        )
        self.assertEqual(
            calculator["K16"].value,
            "Target outcome is not defined pending clinical review.",
        )
        self.assertIn("Provisional pre-cachexia candidate", calculator["G20"].value)
        self.assertNotIn("Option B", calculator["G20"].value)

    def test_named_assumptions_match_canonical_config(self):
        config = json.loads(
            (ROOT / "config" / "simulation_assumptions.v1.json").read_text()
        )
        sheet = self.workbook["Assumptions"]
        values = {
            sheet.cell(row, 1).value: sheet.cell(row, 2).value
            for row in range(8, sheet.max_row + 1)
        }
        self.assertEqual(
            values["PreLower"],
            config["definitions"]["precachexia_lower_weight_loss_percent_exclusive"],
        )
        self.assertEqual(
            values["Category6Intercept"],
            config["illustrative_category_model"]["six_month"]["intercept"],
        )

    def test_clinical_facing_outputs_export_no_numeric_category_value(self):
        calculator = self.workbook["Calculator"]
        results = self.workbook["Results"]
        self.assertEqual(calculator["K9"].number_format, "General")
        self.assertEqual(calculator["K13"].number_format, "General")
        self.assertEqual(results["B24"].value, "illustrative simulation category")
        self.assertIn(
            "Target outcome is not defined pending clinical review.",
            results["B26"].value,
        )
        self.assertNotIn("EXP(", results["B25"].value)
        self.assertNotIn("medium", results["B25"].value)

    def test_workbook_is_macro_free(self):
        self.assertFalse(self.workbook.vba_archive)
        self.assertEqual(WORKBOOK.suffix, ".xlsx")


if __name__ == "__main__":
    unittest.main()
