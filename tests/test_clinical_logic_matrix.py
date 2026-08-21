from __future__ import annotations

import itertools
import json
import unittest
from pathlib import Path

from openpyxl import load_workbook

from cachexia_poc.outcomes import (
    DEFAULT_PRECACHEXIA_CONFIG,
    _fearon_status,
    _precachexia_status,
)
from scripts.build_clinical_logic_review import (
    APPETITE_STATES,
    BMI_STATES,
    LOSS_STATES,
    SARCOPENIA_STATES,
)

ROOT = Path(__file__).resolve().parents[1]
MATRIX_PATH = ROOT / "data" / "clinical_logic_matrix.v1.json"
WORKBOOK_PATH = ROOT / "excel" / "clinical_logic_review_matrix.v1.xlsx"


class ClinicalLogicMatrixTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.payload = json.loads(MATRIX_PATH.read_text(encoding="utf-8"))
        cls.cases = cls.payload["cases"]
        cls.workbook = load_workbook(WORKBOOK_PATH, data_only=False)

    def test_matrix_contains_every_cartesian_combination_once(self):
        expected = set(
            itertools.product(
                (name for name, _ in LOSS_STATES),
                (name for name, _ in BMI_STATES),
                SARCOPENIA_STATES,
                APPETITE_STATES,
            )
        )
        actual = {
            (
                case["weight_loss_state"],
                case["bmi_state"],
                case["sarcopenia"],
                case["reduced_appetite"],
            )
            for case in self.cases
        }
        self.assertEqual(len(self.cases), 324)
        self.assertEqual(len({case["case_id"] for case in self.cases}), 324)
        self.assertEqual(actual, expected)
        self.assertEqual(self.payload["metadata"]["case_count"], 324)

    def test_all_matrix_expectations_match_python_production_logic(self):
        for case in self.cases:
            with self.subTest(case=case["case_id"]):
                loss = case["weight_loss_percent"]
                if loss is None:
                    cachexia, early_risk = "unknown", "unknown"
                else:
                    cachexia, _ = _fearon_status(
                        loss,
                        case["bmi"],
                        case["sarcopenia"],
                    )
                    early_risk, _ = _precachexia_status(
                        loss,
                        case["reduced_appetite"],
                        cachexia,
                        DEFAULT_PRECACHEXIA_CONFIG,
                    )
                self.assertEqual(cachexia, case["expected_cachexia"])
                self.assertEqual(
                    early_risk,
                    case["expected_early_risk"],
                )

    def test_review_workbook_contains_all_cases_and_decision_controls(self):
        self.assertEqual(
            self.workbook.sheetnames,
            [
                "START HERE",
                "Key Scenarios",
                "Review Decisions",
                "Full Logic Matrix",
                "Category Assumptions",
            ],
        )
        start = self.workbook["START HERE"]
        self.assertEqual(
            start["A2"].value,
            "Cachexia and provisional early-risk classification logic.",
        )
        start_text = " ".join(
            str(cell.value)
            for row in start.iter_rows()
            for cell in row
            if cell.value is not None
        )
        self.assertIn(">1% lower bound has no consensus basis", start_text)
        self.assertIn("Eligibility and inclusion criteria are not defined", start_text)
        self.assertIn("10.1016/S1470-2045(10)70218-7", start_text)
        scenarios = self.workbook["Key Scenarios"]
        self.assertEqual(scenarios.max_row, 19)
        self.assertEqual(scenarios["A8"].value, "No weight loss")
        self.assertIn("Loss over 5%", [scenarios.cell(row, 1).value for row in range(8, 20)])
        self.assertIn("J8:J19", str(list(scenarios.data_validations.dataValidation)[0].sqref))
        self.assertEqual(scenarios["J7"].value, "Review status")
        self.assertEqual(
            scenarios["K7"].value,
            "Clinical comments or suggested revisions",
        )

        decisions = self.workbook["Review Decisions"]
        self.assertEqual(decisions["B7"].value, "Clinical question")
        self.assertIn("greater than 5%", decisions["B8"].value)
        decision_text = " ".join(
            str(cell.value)
            for row in decisions.iter_rows()
            for cell in row
            if cell.value is not None
        )
        self.assertIn("eligibility and inclusion criteria", decision_text)
        self.assertIn("Baseline appetite is never carried forward", decision_text)
        self.assertIn("target outcome or estimand", decision_text)

        matrix = self.workbook["Full Logic Matrix"]
        self.assertIn("NOT FOR CLINICAL USE", matrix["A4"].value)
        self.assertEqual(matrix.max_row, 331)
        self.assertEqual(matrix["A8"].value, "LOGIC-001")
        self.assertEqual(matrix["A331"].value, "LOGIC-324")
        validations = list(matrix.data_validations.dataValidation)
        self.assertEqual(len(validations), 1)
        self.assertIn("agree", validations[0].formula1)
        self.assertIn("L8:L331", str(validations[0].sqref))

    def test_category_terms_are_labelled_as_unvalidated_assumptions(self):
        sheet = self.workbook["Category Assumptions"]
        statuses = {
            sheet.cell(row, 5).value
            for row in range(8, sheet.max_row + 1)
            if isinstance(sheet.cell(row, 4).value, (int, float))
            and sheet.cell(row, 5).value
        }
        self.assertTrue(statuses)
        self.assertTrue(
            all(
                "not clinically validated" in status
                or "not a clinical" in status
                for status in statuses
            )
        )
        sheet_text = " ".join(
            str(cell.value)
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
        self.assertIn(
            "Target outcome is not defined pending clinical review",
            sheet_text,
        )

    def test_workbook_contains_no_named_reviewers(self):
        blocked = ("Mari" + "ana", "Nami" + "tha")
        for sheet in self.workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        self.assertFalse(
                            any(name.casefold() in cell.value.casefold() for name in blocked),
                            f"Named reviewer found in {sheet.title}!{cell.coordinate}",
                        )

    def test_workbook_uses_neutral_review_language(self):
        disallowed = ("plain-language", "your review")
        for sheet in self.workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        folded = cell.value.casefold()
                        self.assertFalse(
                            any(phrase in folded for phrase in disallowed),
                            f"Non-neutral wording found in {sheet.title}!{cell.coordinate}",
                        )

    def test_all_workbook_cells_use_times_new_roman(self):
        for sheet in self.workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        self.assertEqual(
                            cell.font.name,
                            "Times New Roman",
                            f"Unexpected font in {sheet.title}!{cell.coordinate}",
                        )
                        self.assertIsNone(
                            cell.font.scheme,
                            f"Theme font overrides Times New Roman in {sheet.title}!{cell.coordinate}",
                        )
                        self.assertEqual(
                            cell.font.family,
                            1.0,
                            f"Unexpected font family in {sheet.title}!{cell.coordinate}",
                        )
        self.assertEqual(
            self.workbook["Key Scenarios"]["K8"].font.name,
            "Times New Roman",
        )
        self.assertEqual(
            self.workbook["Review Decisions"]["E8"].font.name,
            "Times New Roman",
        )

    def test_non_table_blank_cells_have_no_fill_or_direct_style(self):
        for sheet_name, references in {
            "START HERE": ("A3", "B3", "P6"),
            "Key Scenarios": ("A3", "B3", "P6"),
            "Review Decisions": ("A3", "B3", "P6"),
            "Full Logic Matrix": ("A3", "B3", "P6"),
            "Category Assumptions": ("A3", "B3", "P6"),
        }.items():
            sheet = self.workbook[sheet_name]
            for reference in references:
                cell = sheet[reference]
                self.assertIsNone(
                    cell.fill.fill_type,
                    f"Unexpected fill in {sheet_name}!{reference}",
                )
                self.assertEqual(
                    cell.style_id,
                    0,
                    f"Unexpected direct style in {sheet_name}!{reference}",
                )


if __name__ == "__main__":
    unittest.main()
