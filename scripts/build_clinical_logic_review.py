#!/usr/bin/env python3
"""Generate exhaustive classification cases and a clinician review workbook."""

from __future__ import annotations

import itertools
import json
from collections import Counter
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config" / "simulation_assumptions.v1.json"
JSON_OUTPUT = ROOT / "data" / "clinical_logic_matrix.v1.json"
WORKBOOK_OUTPUT = ROOT / "excel" / "clinical_logic_review_matrix.v1.xlsx"

LOSS_STATES = (
    ("unavailable", None),
    ("gain_minus_2", -2.0),
    ("stable_0", 0.0),
    ("pre_lower_exact_1", 1.0),
    ("limited_1_5", 1.5),
    ("fearon_conditional_exact_2", 2.0),
    ("conditional_3", 3.0),
    ("primary_exact_5", 5.0),
    ("primary_over_5", 5.1),
)
BMI_STATES = (
    ("unavailable", None),
    ("below_20", 19.9),
    ("exact_20", 20.0),
    ("above_20", 25.0),
)
SARCOPENIA_STATES = ("yes", "no", "unknown")
APPETITE_STATES = ("yes", "no", "unknown")

NAVY = "17324D"
TEAL = "006D71"
PALE_TEAL = "DDEFEF"
PALE_ORANGE = "FFF1E9"
WHITE = "FFFFFF"
GREY = "E7ECEE"
PALE_GREEN = "E2F0D9"
PALE_RED = "FCE4D6"
PALE_YELLOW = "FFF2CC"
THIN = Side(style="thin", color="B7C4C8")

LOSS_LABELS = {
    "unavailable": "Not available",
    "gain_minus_2": "2% weight gain",
    "stable_0": "No weight change",
    "pre_lower_exact_1": "Exactly 1% loss",
    "limited_1_5": "1.5% loss",
    "fearon_conditional_exact_2": "Exactly 2% loss",
    "conditional_3": "3% loss",
    "primary_exact_5": "Exactly 5% loss",
    "primary_over_5": "5.1% loss",
}

BMI_LABELS = {
    "unavailable": "Not available",
    "below_20": "19.9 (below 20)",
    "exact_20": "20.0 (exact boundary)",
    "above_20": "25.0 (above 20)",
}


def reference_cachexia(
    loss_percent: float | None,
    bmi: float | None,
    sarcopenia: str,
    definitions: dict[str, Any],
) -> tuple[str, str]:
    """Independent decision-table interpretation of the configured branches."""

    if loss_percent is None:
        return "unknown", "Weight loss is unavailable."
    if loss_percent > definitions["fearon_weight_loss_primary_exclusive"]:
        return "yes", "Weight loss is >5%."
    if loss_percent <= definitions["fearon_weight_loss_conditional_exclusive"]:
        return "no", "Loss is <=2%; both conditional branches require >2%."
    if bmi is not None and bmi < definitions["fearon_bmi_exclusive"]:
        return "yes", "Weight loss is >2% and BMI is <20 kg/m²."
    if (
        definitions["fearon_sarcopenia_branch_enabled"]
        and sarcopenia == "yes"
    ):
        return "yes", "Weight loss is >2% and sarcopenia is documented."
    if (
        bmi is not None
        and bmi >= definitions["fearon_bmi_exclusive"]
        and (
            not definitions["fearon_sarcopenia_branch_enabled"]
            or sarcopenia == "no"
        )
    ):
        return "no", "BMI and sarcopenia conditional branches are refuted."
    unavailable = []
    if bmi is None:
        unavailable.append("BMI")
    if (
        definitions["fearon_sarcopenia_branch_enabled"]
        and sarcopenia == "unknown"
    ):
        unavailable.append("sarcopenia")
    return "unknown", "Not evaluable because " + " and ".join(unavailable) + " are unknown."


def reference_early_risk(
    loss_percent: float | None,
    appetite: str,
    cachexia: str,
    definitions: dict[str, Any],
) -> tuple[str, str]:
    """Independent interpretation of the provisional early-risk rule."""

    if cachexia == "yes":
        return "no", "Cachexia criteria take precedence."
    if cachexia == "unknown":
        return "unknown", "Cachexia has not been excluded."
    if loss_percent is None:
        return "unknown", "Weight loss is unavailable."
    lower = definitions["precachexia_lower_weight_loss_percent_exclusive"]
    upper = definitions["precachexia_upper_weight_loss_percent_inclusive"]
    if not lower < loss_percent <= upper:
        return "no", f"Weight loss is outside ({lower}, {upper}]."
    if appetite == "yes":
        return "yes", "Limited loss and reduced appetite are both present."
    if appetite == "no":
        return "no", "Reduced appetite is explicitly absent."
    return "unknown", "Reduced appetite is unknown."


def generate_matrix(config: dict[str, Any]) -> list[dict[str, Any]]:
    definitions = config["definitions"]
    cases = []
    combinations = itertools.product(
        LOSS_STATES,
        BMI_STATES,
        SARCOPENIA_STATES,
        APPETITE_STATES,
    )
    for index, ((loss_state, loss), (bmi_state, bmi), sarcopenia, appetite) in enumerate(
        combinations, 1
    ):
        cachexia, cachexia_reason = reference_cachexia(
            loss, bmi, sarcopenia, definitions
        )
        early_risk, early_risk_reason = reference_early_risk(
            loss, appetite, cachexia, definitions
        )
        boundary = (
            loss in {None, 1.0, 2.0, 5.0}
            or bmi in {None, 20.0}
            or sarcopenia == "unknown"
            or appetite == "unknown"
        )
        cases.append(
            {
                "case_id": f"LOGIC-{index:03d}",
                "weight_loss_state": loss_state,
                "weight_loss_percent": loss,
                "bmi_state": bmi_state,
                "bmi": bmi,
                "sarcopenia": sarcopenia,
                "reduced_appetite": appetite,
                "expected_cachexia": cachexia,
                "cachexia_rationale": cachexia_reason,
                "expected_early_risk": early_risk,
                "early_risk_rationale": early_risk_reason,
                "boundary_or_missing_case": boundary,
                "review_status": "pending",
            }
        )
    return cases


def write_json(config: dict[str, Any], cases: list[dict[str, Any]]) -> None:
    payload = {
        "metadata": {
            "version": "1.0.0",
            "purpose": "Exhaustive software-conformance and clinician-review matrix",
            "warning": (
                "Synthetic rule combinations only. Agreement does not establish "
                "clinical validity or approval."
            ),
            "clinical_approval": "pending clinical review",
            "combination_scope": (
                "9 weight-loss states x 4 BMI states x 3 sarcopenia states "
                "x 3 appetite states"
            ),
            "case_count": len(cases),
            "config_version": config["metadata"]["config_version"],
        },
        "states": {
            "weight_loss": [
                {"name": name, "value": value} for name, value in LOSS_STATES
            ],
            "bmi": [{"name": name, "value": value} for name, value in BMI_STATES],
            "sarcopenia": list(SARCOPENIA_STATES),
            "reduced_appetite": list(APPETITE_STATES),
        },
        "cases": cases,
    }
    JSON_OUTPUT.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def add_title(sheet, title: str, subtitle: str) -> None:
    sheet.merge_cells("A1:P1")
    sheet["A1"] = title
    sheet["A1"].font = Font(size=20, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32
    sheet.merge_cells("A2:P2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(italic=True, color="44545B")
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")


def add_warning(sheet) -> None:
    sheet.merge_cells("A4:P5")
    sheet["A4"] = (
        "NOT FOR CLINICAL USE. These are synthetic rule combinations for logic "
        "review. Agreement confirms the intended implementation only; it does "
        "not establish clinical validity, accuracy, or approval."
    )
    sheet["A4"].font = Font(bold=True, color="7A2F12")
    sheet["A4"].fill = PatternFill("solid", fgColor=PALE_ORANGE)
    sheet["A4"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet["A4"].border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_start_sheet(workbook: Workbook, cases: list[dict[str, Any]]) -> None:
    sheet = workbook.active
    sheet.title = "START HERE"
    add_title(
        sheet,
        "Clinical classification logic review",
        "Cachexia and provisional early-risk classification logic.",
    )
    add_warning(sheet)
    counts = Counter(
        (case["expected_cachexia"], case["expected_early_risk"]) for case in cases
    )
    instructions = [
        ("Suggested review sequence", "Key Scenarios presents 12 representative examples, followed by the main clinical questions in Review Decisions."),
        ("Recording scenario feedback", "Each scenario includes a review-status dropdown and a field for comments or suggested revisions."),
        ("Optional detailed reference", "Full Logic Matrix contains all 324 combinations for focused checking of boundaries and unknown values; row-by-row review is not expected."),
        ("Cachexia rule", "Yes when loss is >5%; or loss is >2% with BMI <20; or loss is >2% with explicitly documented sarcopenia."),
        ("Provisional early-risk rule", "Yes only when cachexia is excluded, loss is >1% and <=5%, and reduced appetite is yes."),
        ("Result values", "Yes = criteria met; no = criteria not met; unknown = available information cannot confirm or exclude the criteria."),
        ("Meaning of unknown", "Unknown is not treated as no. A result remains unknown when the available information cannot confirm or exclude the rule."),
        ("Meaning of sarcopenia", "Yes = independently documented; no = assessed and absent; unknown = not assessed or not documented. It is never inferred."),
        ("Important assumption", "These test cases treat measured weight loss as involuntary because there is no separate involuntary-loss field."),
        ("Risk percentages", "Risk Assumptions contains illustrative simulation terms only. They are not calibrated probabilities or validated clinical effects."),
        ("Current status", "All rules remain pending clinical review. Recorded feedback documents the review outcome but does not validate the prototype."),
    ]
    sheet["A8"], sheet["B8"] = "Topic", "Current interpretation"
    for cell in (sheet["A8"], sheet["B8"]):
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
    for row, (topic, text) in enumerate(instructions, 9):
        sheet.cell(row, 1, topic)
        sheet.cell(row, 2, text)
    sheet["A21"] = "Full-matrix output summary"
    sheet["A22"], sheet["B22"], sheet["C22"] = (
        "Cachexia",
        "Early-risk pattern",
        "Case count",
    )
    for row, ((cachexia, early_risk), count) in enumerate(sorted(counts.items()), 23):
        sheet.cell(row, 1, cachexia)
        sheet.cell(row, 2, early_risk)
        sheet.cell(row, 3, count)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 100
    sheet.column_dimensions["C"].width = 16
    for row in range(8, sheet.max_row + 1):
        sheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")


def add_result_formatting(sheet, ranges: tuple[str, ...]) -> None:
    for cell_range in ranges:
        sheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=['"yes"'],
                fill=PatternFill("solid", fgColor=PALE_GREEN),
            ),
        )
        sheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=['"no"'],
                fill=PatternFill("solid", fgColor=PALE_RED),
            ),
        )
        sheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=['"unknown"'],
                fill=PatternFill("solid", fgColor=PALE_YELLOW),
            ),
        )


def add_review_validation(sheet, cell_range: str) -> None:
    decision = DataValidation(
        type="list",
        formula1='"pending,agree,disagree,question"',
        allow_blank=False,
    )
    decision.errorStyle = "stop"
    decision.error = "Select pending, agree, disagree, or question."
    decision.showErrorMessage = True
    decision.promptTitle = "Review status"
    decision.prompt = "Available values: pending, agree, disagree, or question."
    decision.showInputMessage = True
    sheet.add_data_validation(decision)
    decision.add(cell_range)


def build_key_scenarios_sheet(
    workbook: Workbook, cases: list[dict[str, Any]]
) -> None:
    sheet = workbook.create_sheet("Key Scenarios")
    add_title(
        sheet,
        "Key scenarios to review first",
        "Representative examples covering normal cases, thresholds, and missing information.",
    )
    add_warning(sheet)
    scenario_specs = [
        ("No weight loss", "stable_0", "above_20", "no", "no"),
        ("Limited loss with reduced appetite", "limited_1_5", "above_20", "no", "yes"),
        ("Limited loss without reduced appetite", "limited_1_5", "above_20", "no", "no"),
        ("Exactly 1% loss", "pre_lower_exact_1", "above_20", "no", "yes"),
        ("Exactly 2% loss", "fearon_conditional_exact_2", "below_20", "yes", "yes"),
        ("Loss over 2% with low BMI", "conditional_3", "below_20", "no", "no"),
        ("Loss over 2% with sarcopenia", "conditional_3", "above_20", "yes", "no"),
        ("Loss over 2%; sarcopenia unknown", "conditional_3", "above_20", "unknown", "no"),
        ("Exactly 5% loss with reduced appetite", "primary_exact_5", "above_20", "no", "yes"),
        ("Loss over 5%", "primary_over_5", "above_20", "no", "no"),
        ("Weight loss unavailable", "unavailable", "above_20", "no", "yes"),
        ("BMI unavailable; sarcopenia absent", "conditional_3", "unavailable", "no", "yes"),
    ]
    case_lookup = {
        (
            case["weight_loss_state"],
            case["bmi_state"],
            case["sarcopenia"],
            case["reduced_appetite"],
        ): case
        for case in cases
    }
    headers = [
        "Scenario",
        "Weight change",
        "BMI",
        "Sarcopenia",
        "Reduced appetite",
        "Cachexia result",
        "Reason for result",
        "Early-risk result",
        "Reason for result",
        "Review status",
        "Clinical comments or suggested revisions",
    ]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(7, column, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row, spec in enumerate(scenario_specs, 8):
        title, loss_state, bmi_state, sarcopenia, appetite = spec
        case = case_lookup[(loss_state, bmi_state, sarcopenia, appetite)]
        values = [
            title,
            LOSS_LABELS[loss_state],
            BMI_LABELS[bmi_state],
            sarcopenia,
            appetite,
            case["expected_cachexia"],
            case["cachexia_rationale"],
            case["expected_early_risk"],
            case["early_risk_rationale"],
            "pending",
            None,
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row, column, value)
        sheet.cell(row, 10).fill = PatternFill("solid", fgColor=PALE_TEAL)
        sheet.cell(row, 11).fill = PatternFill("solid", fgColor=PALE_TEAL)
        sheet.row_dimensions[row].height = 48
    last_row = 7 + len(scenario_specs)
    add_review_validation(sheet, f"J8:J{last_row}")
    add_result_formatting(sheet, (f"F8:F{last_row}", f"H8:H{last_row}"))
    sheet.auto_filter.ref = f"A7:K{last_row}"
    sheet.freeze_panes = "A8"
    widths = (34, 24, 24, 18, 20, 18, 48, 22, 48, 18, 48)
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width


def build_matrix_sheet(workbook: Workbook, cases: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Full Logic Matrix")
    add_title(
        sheet,
        "Complete logic matrix",
        "All 324 combinations, available for filtered review of thresholds and missing-value patterns.",
    )
    add_warning(sheet)
    headers = [
        "Case ID",
        "Weight change",
        "Loss %",
        "BMI",
        "Sarcopenia",
        "Reduced appetite",
        "Cachexia result",
        "Reason for result",
        "Early-risk result",
        "Reason for result",
        "Boundary/missing?",
        "Review status",
        "Clinical comments or suggested revisions",
    ]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(7, column, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row, case in enumerate(cases, 8):
        values = [
            case["case_id"],
            LOSS_LABELS[case["weight_loss_state"]],
            case["weight_loss_percent"],
            BMI_LABELS[case["bmi_state"]],
            case["sarcopenia"],
            case["reduced_appetite"],
            case["expected_cachexia"],
            case["cachexia_rationale"],
            case["expected_early_risk"],
            case["early_risk_rationale"],
            "yes" if case["boundary_or_missing_case"] else "no",
            "pending",
            None,
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row, column, value)
        for column in (12, 13):
            sheet.cell(row, column).fill = PatternFill("solid", fgColor=PALE_TEAL)
    last_row = 7 + len(cases)
    add_review_validation(sheet, f"L8:L{last_row}")
    add_result_formatting(sheet, (f"G8:G{last_row}", f"I8:I{last_row}"))
    sheet.auto_filter.ref = f"A7:M{last_row}"
    sheet.freeze_panes = "A8"
    widths = (14, 25, 12, 24, 16, 18, 18, 52, 22, 52, 18, 18, 55)
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in range(8, last_row + 1):
        for column in (8, 10, 13):
            sheet.cell(row, column).alignment = Alignment(
                wrap_text=True, vertical="top"
            )


def build_risk_terms_sheet(workbook: Workbook, config: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Risk Assumptions")
    add_title(
        sheet,
        "Illustrative risk-score terms",
        "Simulation assumptions only; these values are not validated clinical effects.",
    )
    add_warning(sheet)
    headers = ["Horizon", "Term", "Category", "Value", "Status"]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(7, column, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
    row = 8
    for horizon in ("three_month", "six_month"):
        terms = config["risk_outputs"][horizon]
        scalar_terms = {
            "intercept": terms["intercept"],
            "age_over_55": terms["age_over_55"],
            "baseline_weight_loss_per_percent": terms[
                "baseline_weight_loss_per_percent"
            ],
            "low_bmi_under_20": terms["low_bmi_under_20"],
            "cancer_type_multiplier": terms["cancer_type_multiplier"],
        }
        for term, value in scalar_terms.items():
            sheet.append(
                [
                    horizon,
                    term,
                    "all",
                    value,
                    "simulation assumption; not clinically validated",
                ]
            )
            row += 1
        for term in ("stage", "ecog", "appetite"):
            for category, value in terms[term].items():
                sheet.append(
                    [
                        horizon,
                        term,
                        category,
                        value,
                        "simulation assumption; not clinically validated",
                    ]
                )
                row += 1
    sheet.append([])
    sheet.append(
        ["Shared", "cancer risk multiplier", "Cancer type", "Value", "Status"]
    )
    for cancer_type, value in config["simulation_relationships"][
        "cancer_risk_multipliers"
    ].items():
        sheet.append(
            [
                "both",
                "cancer risk multiplier",
                cancer_type,
                value,
                "supplied illustrative assumption; not a relative risk",
            ]
        )
    sheet.freeze_panes = "A8"
    sheet.auto_filter.ref = f"A7:E{sheet.max_row}"
    for column, width in enumerate((18, 38, 28, 16, 55), 1):
        sheet.column_dimensions[get_column_letter(column)].width = width


def build_decisions_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Review Decisions")
    add_title(
        sheet,
        "Clinical decision record",
        "Pending indicates that no clinical decision has been recorded.",
    )
    add_warning(sheet)
    headers = [
        "Decision ID",
        "Clinical question",
        "Current prototype rule",
        "Decision",
        "Suggested revision",
        "Rationale or comments",
        "Reviewed by / role",
        "Date",
    ]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(7, column, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    topics = [
        ("MATRIX-001", "Should the primary cachexia threshold require loss greater than 5%, excluding exactly 5%?", "Weight loss >5%"),
        ("MATRIX-002", "Should the low-BMI branch require both loss greater than 2% and BMI below 20?", "Weight loss >2% and BMI <20"),
        ("MATRIX-003", "Should documented sarcopenia identify cachexia when weight loss is greater than 2%?", "Weight loss >2% and documented sarcopenia"),
        ("MATRIX-004", "Should a classification remain unknown when available evidence cannot confirm or exclude a criterion?", "Unknown is kept distinct from no"),
        ("MATRIX-005", "Is the proposed early-risk weight-loss interval appropriate?", "Weight loss >1% and <=5%, after cachexia is excluded"),
        ("MATRIX-006", "Should reduced appetite be required for the proposed early-risk pattern?", "Reduced appetite = yes"),
        ("MATRIX-007", "What evidence should count as documented sarcopenia, and can baseline evidence be carried into future outcome labels?", "Explicit yes/no/unknown evidence; provisional baseline carry-forward"),
        ("MATRIX-008", "Should involuntary weight loss be recorded separately rather than assumed for synthetic rule cases?", "Measured loss is treated as involuntary in rule-testing scenarios"),
        ("MATRIX-009", "Should the illustrative risk percentages remain in the demonstrator?", "Configured simulation coefficients and multipliers; not clinically validated"),
    ]
    for row, values in enumerate(topics, 8):
        sheet.cell(row, 1, values[0])
        sheet.cell(row, 2, values[1])
        sheet.cell(row, 3, values[2])
        sheet.cell(row, 4, "pending")
        for column in range(4, 9):
            sheet.cell(row, column).fill = PatternFill("solid", fgColor=PALE_TEAL)
        sheet.row_dimensions[row].height = 44
    decision = DataValidation(
        type="list",
        formula1='"pending,approved,rejected,revision_requested"',
        allow_blank=False,
    )
    decision.errorStyle = "stop"
    decision.error = "Select pending, approved, rejected, or revision_requested."
    decision.showErrorMessage = True
    decision.promptTitle = "Decision status"
    decision.prompt = "Available values: pending, approved, rejected, or revision_requested."
    decision.showInputMessage = True
    sheet.add_data_validation(decision)
    decision.add(f"D8:D{7 + len(topics)}")
    widths = (16, 30, 55, 22, 60, 50, 24, 18)
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in range(8, 8 + len(topics)):
        for column in range(3, 7):
            sheet.cell(row, column).alignment = Alignment(
                wrap_text=True, vertical="top"
            )


def style_workbook(workbook: Workbook) -> None:
    for named_style in workbook._named_styles:
        if named_style.name == "Normal":
            named_style.font = Font(name="Times New Roman", size=11)
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.sheet_view.zoomScale = 90
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        if sheet.title in {"Key Scenarios", "Full Logic Matrix", "Review Decisions"}:
            sheet.page_setup.orientation = "landscape"
        for row in sheet.iter_rows():
            for cell in row:
                font = copy(cell.font)
                font.name = "Times New Roman"
                cell.font = font
                if cell.value is not None and cell.row >= 7:
                    cell.alignment = Alignment(
                        wrap_text=cell.alignment.wrap_text,
                        vertical=cell.alignment.vertical or "top",
                    )


def main() -> None:
    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    cases = generate_matrix(config)
    expected_count = (
        len(LOSS_STATES)
        * len(BMI_STATES)
        * len(SARCOPENIA_STATES)
        * len(APPETITE_STATES)
    )
    if len(cases) != expected_count:
        raise RuntimeError("Classification matrix is incomplete.")
    JSON_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    WORKBOOK_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    write_json(config, cases)
    workbook = Workbook()
    build_start_sheet(workbook, cases)
    build_key_scenarios_sheet(workbook, cases)
    build_decisions_sheet(workbook)
    build_matrix_sheet(workbook, cases)
    build_risk_terms_sheet(workbook, config)
    style_workbook(workbook)
    workbook.active = 0
    workbook.save(WORKBOOK_OUTPUT)
    print(
        f"Built {JSON_OUTPUT.relative_to(ROOT)} and "
        f"{WORKBOOK_OUTPUT.relative_to(ROOT)} with {len(cases)} cases"
    )


if __name__ == "__main__":
    main()
