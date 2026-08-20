#!/usr/bin/env python3
"""Build the macro-free interactive Excel prototype from canonical artifacts."""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "excel" / "cachexia_risk_prototype.v1.xlsx"
CONFIG_PATH = ROOT / "config" / "simulation_assumptions.v1.json"
DATA_PATH = ROOT / "data" / "synthetic_patients.v1.json"

NAVY = "17324D"
TEAL = "006D71"
PALE_TEAL = "DDEFEF"
PALE_ORANGE = "FFF1E9"
ORANGE = "C55A11"
WHITE = "FFFFFF"
GREY = "E7ECEE"
RED = "C00000"
GREEN = "548235"
THIN = Side(style="thin", color="B7C4C8")


def title(sheet, text: str, subtitle: str | None = None) -> None:
    sheet.merge_cells("A1:H1")
    cell = sheet["A1"]
    cell.value = text
    cell.font = Font(size=20, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32
    if subtitle:
        sheet.merge_cells("A2:H2")
        sheet["A2"] = subtitle
        sheet["A2"].font = Font(italic=True, color="44545B")
        sheet["A2"].alignment = Alignment(wrap_text=True)
        sheet.row_dimensions[2].height = 34


def warning(sheet, row: int) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=8)
    cell = sheet.cell(row, 1)
    cell.value = (
        "NOT FOR CLINICAL USE. All patients and outputs are synthetic. Risk "
        "relationships are simulation assumptions, not clinically validated "
        "effects. Do not use for diagnosis, prognosis, treatment, patient "
        "care, or medical decisions."
    )
    cell.font = Font(bold=True, color="7A2F12")
    cell.fill = PatternFill("solid", fgColor=PALE_ORANGE)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    sheet.row_dimensions[row].height = 34


def add_list_validation(sheet, cell: str, values: list[str]) -> None:
    formula = '"' + ",".join(values) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=False)
    validation.error = "Choose a value from the list; unknown is distinct from no."
    validation.errorTitle = "Invalid category"
    validation.prompt = "Select a permitted value."
    validation.promptTitle = "Validated input"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(sheet[cell])


def define_name(workbook: Workbook, name: str, reference: str) -> None:
    workbook.defined_names.add(DefinedName(name, attr_text=reference))


def build_assumptions(workbook: Workbook, config: dict) -> None:
    sheet = workbook.create_sheet("Assumptions")
    title(
        sheet,
        "Central simulation assumptions",
        "Editable for research scenarios only. Values are not validated clinical effects.",
    )
    warning(sheet, 4)
    sheet.append([])
    sheet.append(["Key", "Value", "Status / source"])
    scalar_rows = [
        ("DaysPerMonth", config["definitions"]["days_per_month"], "Operational normalization"),
        ("TrajectoryEpsilon", config["definitions"]["trajectory_epsilon_percent"], "Simulation assumption"),
        ("FearonPrimary", config["definitions"]["fearon_weight_loss_primary_exclusive"], "Implemented criterion boundary"),
        ("FearonConditional", config["definitions"]["fearon_weight_loss_conditional_exclusive"], "Implemented criterion boundary"),
        ("FearonBMI", config["definitions"]["fearon_bmi_exclusive"], "Implemented criterion boundary"),
        ("PreLower", config["definitions"]["precachexia_lower_weight_loss_percent_exclusive"], "PROVISIONAL — review required"),
        ("PreUpper", config["definitions"]["precachexia_upper_weight_loss_percent_inclusive"], "PROVISIONAL — review required"),
        ("AgeThreshold", config["risk_outputs"]["age_threshold_exclusive"], "Simulation assumption"),
        ("BandLow", config["risk_outputs"]["band_thresholds"]["low_upper_exclusive"], "Simulation assumption"),
        ("BandHigh", config["risk_outputs"]["band_thresholds"]["high_lower_inclusive"], "Simulation assumption"),
        ("WeightMin", config["cohort"]["historical_weight_kg"]["minimum"], "Input validation bound"),
        ("WeightMax", config["cohort"]["historical_weight_kg"]["maximum"], "Input validation bound"),
    ]
    for horizon, prefix in (("three_month", "Risk3"), ("six_month", "Risk6")):
        values = config["risk_outputs"][horizon]
        scalar_rows.extend(
            [
                (f"{prefix}Intercept", values["intercept"], "Simulation assumption"),
                (f"{prefix}Age", values["age_over_55"], "Simulation assumption"),
                (f"{prefix}Loss", values["baseline_weight_loss_per_percent"], "Simulation assumption"),
                (f"{prefix}BMI", values["low_bmi_under_20"], "Simulation assumption"),
                (f"{prefix}Cancer", values["cancer_type_multiplier"], "Simulation assumption"),
            ]
        )
    start = 8
    for offset, (key, value, status) in enumerate(scalar_rows):
        row = start + offset
        sheet.cell(row, 1, key)
        sheet.cell(row, 2, value)
        sheet.cell(row, 3, status)
        define_name(workbook, key, f"'Assumptions'!$B${row}")

    cancer = config["simulation_relationships"]["cancer_latent_points"]
    sheet["E7"], sheet["F7"] = "Cancer type", "Latent points"
    for row, (key, value) in enumerate(cancer.items(), 8):
        sheet.cell(row, 5, key)
        sheet.cell(row, 6, value)
    cancer_end = 7 + len(cancer)

    sheet["H7"], sheet["I7"], sheet["J7"] = "Stage", "3m term", "6m term"
    stages = list(config["risk_outputs"]["three_month"]["stage"])
    for row, key in enumerate(stages, 8):
        sheet.cell(row, 8, key)
        sheet.cell(row, 9, config["risk_outputs"]["three_month"]["stage"][key])
        sheet.cell(row, 10, config["risk_outputs"]["six_month"]["stage"][key])
    stage_end = 7 + len(stages)

    sheet["L7"], sheet["M7"], sheet["N7"] = "ECOG", "3m term", "6m term"
    ecogs = list(config["risk_outputs"]["three_month"]["ecog"])
    for row, key in enumerate(ecogs, 8):
        sheet.cell(row, 12, key)
        sheet.cell(row, 13, config["risk_outputs"]["three_month"]["ecog"][key])
        sheet.cell(row, 14, config["risk_outputs"]["six_month"]["ecog"][key])
    ecog_end = 7 + len(ecogs)

    sheet["P7"], sheet["Q7"], sheet["R7"] = "Appetite", "3m term", "6m term"
    appetites = list(config["risk_outputs"]["three_month"]["appetite"])
    for row, key in enumerate(appetites, 8):
        sheet.cell(row, 16, key)
        sheet.cell(row, 17, config["risk_outputs"]["three_month"]["appetite"][key])
        sheet.cell(row, 18, config["risk_outputs"]["six_month"]["appetite"][key])
    appetite_end = 7 + len(appetites)

    define_name(workbook, "CancerTable", f"'Assumptions'!$E$8:$F${cancer_end}")
    define_name(workbook, "StageTable", f"'Assumptions'!$H$8:$J${stage_end}")
    define_name(workbook, "EcogTable", f"'Assumptions'!$L$8:$N${ecog_end}")
    define_name(workbook, "AppetiteTable", f"'Assumptions'!$P$8:$R${appetite_end}")
    for column in range(1, 19):
        sheet.column_dimensions[get_column_letter(column)].width = 20
    sheet.column_dimensions["C"].width = 32
    sheet.freeze_panes = "A8"


def build_input(workbook: Workbook, config: dict) -> None:
    sheet = workbook.create_sheet("Patient Input")
    title(
        sheet,
        "Interactive synthetic patient input",
        "Change blue cells, then review automatically recalculated outputs on Results.",
    )
    warning(sheet, 4)
    labels = [
        ("Prediction date", date(2026, 1, 31)),
        ("Age (years)", 65),
        ("Sex", "female"),
        ("Cancer type", "lung"),
        ("Cancer subtype", "NSCLC"),
        ("Cancer stage", "III"),
        ("Height (cm; blank = unknown)", 170),
        ("ECOG", "2"),
        ("Reduced appetite", "no"),
        ("Sarcopenia", "unknown"),
    ]
    for row, (label, value) in enumerate(labels, 4):
        target_row = row + 4
        sheet.cell(target_row, 1, label)
        sheet.cell(target_row, 2, value)
        sheet.cell(target_row, 2).fill = PatternFill("solid", fgColor=PALE_TEAL)
        sheet.cell(target_row, 2).border = Border(
            left=THIN, right=THIN, top=THIN, bottom=THIN
        )
    sheet["B8"].number_format = "yyyy-mm-dd"

    add_list_validation(sheet, "B10", ["female", "male", "unknown"])
    add_list_validation(
        sheet, "B11", list(config["cohort"]["cancer_type_probabilities"])
    )
    add_list_validation(sheet, "B13", ["I", "II", "III", "IV", "unknown"])
    add_list_validation(sheet, "B15", ["0", "1", "2", "3", "4", "unknown"])
    add_list_validation(sheet, "B16", ["yes", "no", "unknown"])
    add_list_validation(sheet, "B17", ["yes", "no", "unknown"])
    age = DataValidation(
        type="whole",
        operator="between",
        formula1=str(config["cohort"]["age"]["minimum"]),
        formula2=str(config["cohort"]["age"]["maximum"]),
    )
    age.showErrorMessage = True
    age.errorStyle = "stop"
    age.errorTitle = "Invalid age"
    age.error = "Enter a whole-number age within the configured range."
    height_values = list(config["cohort"]["height_cm"].values())
    height = DataValidation(
        type="decimal",
        operator="between",
        formula1=str(min(item["minimum"] for item in height_values)),
        formula2=str(max(item["maximum"] for item in height_values)),
        allow_blank=True,
    )
    height.showErrorMessage = True
    height.errorStyle = "stop"
    height.errorTitle = "Invalid height"
    height.error = "Enter a height within the configured range or leave it blank."
    sheet.add_data_validation(age)
    sheet.add_data_validation(height)
    age.add(sheet["B9"])
    height.add(sheet["B14"])

    sheet["A20"] = "Dated weight history (predictors use only dates <= prediction date)"
    sheet["A20"].font = Font(size=13, bold=True, color=TEAL)
    sheet["A21"], sheet["B21"], sheet["C21"] = "Measurement date", "Weight (kg)", "Eligibility note"
    sheet["A22"], sheet["B22"] = date(2025, 7, 31), 80
    sheet["A23"], sheet["B23"] = date(2026, 1, 31), 76
    for row in range(22, 34):
        sheet.cell(row, 1).fill = PatternFill("solid", fgColor=PALE_TEAL)
        sheet.cell(row, 2).fill = PatternFill("solid", fgColor=PALE_TEAL)
        sheet.cell(row, 1).number_format = "yyyy-mm-dd"
        sheet.cell(
            row, 3, f'=IF(A{row}="","",IF(A{row}<=$B$8,"baseline-eligible","outcome-only / excluded from predictors"))'
        )
    weight_bounds = config["cohort"]["historical_weight_kg"]
    weight_validation = DataValidation(
        type="decimal",
        operator="between",
        formula1=str(weight_bounds["minimum"]),
        formula2=str(weight_bounds["maximum"]),
        allow_blank=True,
    )
    weight_validation.showErrorMessage = True
    weight_validation.errorStyle = "stop"
    weight_validation.errorTitle = "Invalid weight"
    weight_validation.error = "Enter a weight within the configured range or clear both cells in the row."
    date_validation = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(2000,1,1)",
        formula2="DATE(2100,12,31)",
        allow_blank=True,
    )
    date_validation.showErrorMessage = True
    date_validation.errorStyle = "stop"
    date_validation.errorTitle = "Invalid measurement date"
    date_validation.error = "Enter a valid measurement date or clear both cells in the row."
    sheet.add_data_validation(weight_validation)
    sheet.add_data_validation(date_validation)
    weight_validation.add("B22:B33")
    date_validation.add("A22:A33")
    sheet.column_dimensions["A"].width = 44
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 42
    sheet.freeze_panes = "A8"


def build_results(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Results")
    title(
        sheet,
        "Calculated synthetic outputs",
        "Excel 2019 / Microsoft 365 formulas. Blank means not calculable; unknown is not no.",
    )
    warning(sheet, 4)
    labels = {
        8: "Baseline weight date",
        9: "Baseline/current weight (kg)",
        10: "Prior comparison date",
        11: "Prior comparison weight (kg)",
        12: "BMI (kg/m²)",
        13: "Percentage weight loss (+ = loss)",
        14: "Interval (days)",
        15: "Weight-loss rate (kg/month)",
        16: "Rate (percentage points/month)",
        17: "Trajectory",
        19: "Fearon-supported status",
        20: "Provisional pre-cachexia candidate",
    }
    for row, label in labels.items():
        sheet.cell(row, 1, label)
    dates = "'Patient Input'!$A$22:$A$33"
    weights = "'Patient Input'!$B$22:$B$33"
    sheet["B8"] = (
        f'=IFERROR(AGGREGATE(14,6,{dates}/(({dates}<>"")*'
        f'({dates}<=\'Patient Input\'!$B$8)*({weights}>=WeightMin)*'
        f'({weights}<=WeightMax)),1),"")'
    )
    sheet["B9"] = (
        f'=IF(B8="","",LOOKUP(2,1/(({dates}=B8)*({weights}<>"")*'
        f'({weights}>=WeightMin)*({weights}<=WeightMax)),{weights}))'
    )
    sheet["B10"] = (
        f'=IF(B8="","",IFERROR(AGGREGATE(15,6,{dates}/(({dates}<>"")*'
        f'({dates}>=EDATE(B8,-6))*({dates}<B8)*({weights}>=WeightMin)*'
        f'({weights}<=WeightMax)),1),""))'
    )
    sheet["B11"] = (
        f'=IF(B10="","",LOOKUP(2,1/(({dates}=B10)*({weights}<>"")*'
        f'({weights}>=WeightMin)*({weights}<=WeightMax)),{weights}))'
    )
    sheet["B12"] = '=IF(OR(B9="",\'Patient Input\'!B14=""),"",B9/(\'Patient Input\'!B14/100)^2)'
    sheet["B13"] = '=IF(OR(B9="",B11=""),"",((B11-B9)/B11)*100)'
    sheet["B14"] = '=IF(OR(B8="",B10=""),"",B8-B10)'
    sheet["B15"] = '=IF(OR(B9="",B11="",B14<=0),"",(B11-B9)/(B14/DaysPerMonth))'
    sheet["B16"] = '=IF(OR(B13="",B14<=0),"",B13/(B14/DaysPerMonth))'
    sheet["B17"] = '=IF(B13="","unknown",IF(B13>TrajectoryEpsilon,"loss",IF(B13<-TrajectoryEpsilon,"gain","stable")))'
    sheet["B19"] = (
        '=IF(B13="","unknown",IF(B13>FearonPrimary,"yes",'
        'IF(B13<=FearonConditional,"no",IF(OR(AND(B12<>"",B12<FearonBMI),'
        '\'Patient Input\'!B17="yes"),"yes",IF(AND(B12<>"",B12>=FearonBMI,'
        '\'Patient Input\'!B17="no"),"no","unknown")))))'
    )
    sheet["B20"] = (
        '=IF(B19="unknown","unknown",IF(B19="yes","no",'
        'IF(OR(B13<=PreLower,B13>PreUpper),"no",'
        'IF(\'Patient Input\'!B16="yes","yes",IF(\'Patient Input\'!B16="no","no","unknown")))))'
    )

    sheet["A23"], sheet["B23"], sheet["D23"] = "Simulated horizon", "3 months", "6 months"
    sheet["A24"], sheet["A25"], sheet["A26"], sheet["A27"] = (
        "Simulation score", "Displayed estimate", "Risk band", "Factor explanation"
    )
    common3 = (
        "Risk3Intercept+IF('Patient Input'!B9>AgeThreshold,Risk3Age,0)"
        "+VLOOKUP('Patient Input'!B13,StageTable,2,FALSE)"
        "+VLOOKUP('Patient Input'!B15,EcogTable,2,FALSE)"
        "+VLOOKUP('Patient Input'!B16,AppetiteTable,2,FALSE)"
        "+MAX(0,N(B13))*Risk3Loss+IF(AND(B12<>\"\",B12<FearonBMI),Risk3BMI,0)"
        "+VLOOKUP('Patient Input'!B11,CancerTable,2,FALSE)*Risk3Cancer"
    )
    common6 = (
        "Risk6Intercept+IF('Patient Input'!B9>AgeThreshold,Risk6Age,0)"
        "+VLOOKUP('Patient Input'!B13,StageTable,3,FALSE)"
        "+VLOOKUP('Patient Input'!B15,EcogTable,3,FALSE)"
        "+VLOOKUP('Patient Input'!B16,AppetiteTable,3,FALSE)"
        "+MAX(0,N(B13))*Risk6Loss+IF(AND(B12<>\"\",B12<FearonBMI),Risk6BMI,0)"
        "+VLOOKUP('Patient Input'!B11,CancerTable,2,FALSE)*Risk6Cancer"
    )
    sheet["B24"], sheet["D24"] = (
        f'=IF(OR(B12="",B13=""),"",{common3})',
        f'=IF(OR(B12="",B13=""),"",{common6})',
    )
    sheet["B25"], sheet["D25"] = (
        '=IF(B24="","",1/(1+EXP(-B24)))',
        '=IF(D24="","",1/(1+EXP(-D24)))',
    )
    sheet["B26"] = '=IF(B25="","unknown",IF(B25<BandLow,"low",IF(B25>=BandHigh,"high","medium")))'
    sheet["D26"] = '=IF(D25="","unknown",IF(D25<BandLow,"low",IF(D25>=BandHigh,"high","medium")))'
    factor_formula = (
        'TRIM(IF(\'Patient Input\'!B9>AgeThreshold,"age >55; ","")&'
        'IF(B13>0,"baseline loss "&TEXT(B13,"0.0")&"%; ","")&'
        'IF(AND(B12<>"",B12<FearonBMI),"BMI <20; ","")&'
        '"stage "&\'Patient Input\'!B13&"; ECOG "&\'Patient Input\'!B15&'
        '"; appetite="&\'Patient Input\'!B16&"; "&\'Patient Input\'!B11)'
    )
    withheld = (
        '=IF(OR(B12="",B13=""),'
        '"Estimate withheld: BMI and baseline weight change are required.",'
        f'{factor_formula})'
    )
    sheet["B27"], sheet["D27"] = withheld, withheld
    sheet["A29"] = "Interpretation"
    sheet["B29"] = (
        "These are deterministic simulation outputs, not calibrated probabilities. "
        "Changing six-month assumptions does not change the three-month formula."
    )
    sheet.merge_cells("B29:H30")
    sheet["B29"].alignment = Alignment(wrap_text=True, vertical="top")

    for row in range(8, 28):
        sheet.cell(row, 1).font = Font(bold=True)
        for column in (2, 4):
            sheet.cell(row, column).fill = PatternFill("solid", fgColor=PALE_TEAL)
    for cell in ("B12", "B13", "B15", "B16"):
        sheet[cell].number_format = "0.00"
    for cell in ("B25", "D25"):
        sheet[cell].number_format = "0.0%"
    for cell in ("B8", "B10"):
        sheet[cell].number_format = "yyyy-mm-dd"
    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 34
    sheet.column_dimensions["C"].width = 4
    sheet.column_dimensions["D"].width = 34
    for column in range(5, 9):
        sheet.column_dimensions[get_column_letter(column)].width = 14
    sheet.conditional_formatting.add(
        "B25:D25",
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["BandHigh"],
            fill=PatternFill("solid", fgColor="F4CCCC"),
        ),
    )


def build_cohort(workbook: Workbook, patients: list[dict]) -> None:
    sheet = workbook.create_sheet("Synthetic Cohort")
    title(
        sheet,
        "Generated synthetic cohort",
        "For plausibility review only; counts do not estimate prevalence.",
    )
    warning(sheet, 4)
    headers = [
        "patient_id", "prediction_date", "age", "sex", "cancer_type",
        "cancer_stage", "height_cm", "ecog", "reduced_appetite", "sarcopenia",
        "baseline_bmi", "baseline_loss_percent", "outcome_3m_cachexia",
        "outcome_3m_precachexia", "risk_3m", "outcome_6m_cachexia",
        "outcome_6m_precachexia", "risk_6m", "edge_case",
    ]
    for column, header in enumerate(headers, 1):
        sheet.cell(7, column, header)
    for row, patient in enumerate(patients, 8):
        predictors = patient["baseline_predictors"]
        values = [
            patient["patient_id"], patient["prediction_date"], patient["age"],
            patient["sex"], patient["cancer_type"], patient["cancer_stage"],
            patient["height_cm"], "unknown" if patient["ecog"] is None else patient["ecog"],
            patient["reduced_appetite"], patient["sarcopenia"], predictors["bmi"],
            predictors["weight_loss_percent"], patient["outcome_3m"]["cachexia"],
            patient["outcome_3m"]["precachexia_candidate"],
            patient["simulated_risk_3m"]["probability"],
            patient["outcome_6m"]["cachexia"],
            patient["outcome_6m"]["precachexia_candidate"],
            patient["simulated_risk_6m"]["probability"], patient["edge_case"],
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row, column, value)
    last_row = 7 + len(patients)
    table = Table(displayName="SyntheticCohort", ref=f"A7:S{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    sheet.add_table(table)
    for column in range(1, 20):
        sheet.column_dimensions[get_column_letter(column)].width = 18
    sheet.column_dimensions["A"].width = 24
    sheet.freeze_panes = "A8"
    sheet.auto_filter.ref = f"A7:S{last_row}"
    for row in range(8, last_row + 1):
        sheet.cell(row, 15).number_format = "0.0%"
        sheet.cell(row, 18).number_format = "0.0%"


def build_dictionary(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Data Dictionary")
    title(sheet, "Data dictionary", "Units, values, and unknown behavior.")
    warning(sheet, 4)
    rows = [
        ("patient_id", "text", "SYN-* synthetic identifier"),
        ("prediction_date", "date", "Explicit predictor cutoff"),
        ("age", "years", "18–95"),
        ("sex", "category", "female / male / unknown"),
        ("cancer_type", "category", "Configured solid tumour categories"),
        ("cancer_stage", "category", "I / II / III / IV / unknown"),
        ("height_cm", "cm", "140–200 or blank/unknown"),
        ("weights", "kg + date", "25–160 kg; multiple dated rows"),
        ("ecog", "category", "0–4 / unknown"),
        ("reduced_appetite", "tri-state", "yes / no / unknown"),
        ("sarcopenia", "tri-state", "yes / no / unknown; never inferred"),
        ("outcome_3m", "tri-state labels", "Inclusive 3-calendar-month horizon"),
        ("outcome_6m", "tri-state labels", "Inclusive 6-calendar-month horizon"),
        ("risk outputs", "0–100% display", "Simulation only; not calibrated"),
    ]
    sheet.append([])
    sheet.append([])
    sheet.append([])
    sheet.append(["Field", "Unit/type", "Definition"])
    for row in rows:
        sheet.append(row)
    for column, width in enumerate((28, 22, 78), 1):
        sheet.column_dimensions[get_column_letter(column)].width = width


def build_review(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Clinical Review")
    title(
        sheet,
        "clinical-reviewer and clinical-reviewer review decisions",
        "No approval has been received. Record decisions explicitly; blank is not approval.",
    )
    warning(sheet, 4)
    headers = ["ID", "Question", "Owner", "Status", "Decision", "Rationale", "Effective date"]
    for column, header in enumerate(headers, 1):
        sheet.cell(7, column, header)
    questions = [
        ("CLIN-001", "Confirm/revise predictor ranges and distributions"),
        ("CLIN-002", "Confirm stage/ECOG/appetite simulation relationships"),
        ("CLIN-003", "Approve/reject/revise provisional pre-cachexia rule"),
        ("CLIN-004", "Confirm Fearon implementation and unknown behavior"),
        ("CLIN-005", "Confirm outcome selection and inclusive boundaries"),
        ("CLIN-006", "Confirm sarcopenia representation"),
        ("UX-001", "Identify clinically misleading wording or presentation"),
    ]
    for row, (identifier, question) in enumerate(questions, 8):
        sheet.cell(row, 1, identifier)
        sheet.cell(row, 2, question)
        sheet.cell(row, 3, "clinical-reviewer; clinical-reviewer")
        sheet.cell(row, 4, "pending")
    status = DataValidation(
        type="list",
        formula1='"pending,approved,rejected,revision_requested"',
        allow_blank=False,
    )
    sheet.add_data_validation(status)
    status.add(f"D8:D{7 + len(questions)}")
    widths = (14, 62, 24, 22, 42, 52, 18)
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A8"


def build_readme(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "START HERE"
    title(
        sheet,
        "Synthetic cachexia risk workbook",
        "Interactive, macro-free clinical review prototype.",
    )
    warning(sheet, 4)
    instructions = [
        ("1", "Open Patient Input and change only blue cells."),
        ("2", "Enter an explicit prediction date and dated weight history."),
        ("3", "Open Results to inspect derived variables, transparent labels, and separate simulated horizons."),
        ("4", "Review Assumptions before interpreting any output; every relationship is provisional."),
        ("5", "Use Clinical Review to record clinical-reviewer and clinical-reviewer's decisions."),
    ]
    sheet["A8"], sheet["B8"] = "Step", "Action"
    for row, values in enumerate(instructions, 9):
        sheet.cell(row, 1, values[0])
        sheet.cell(row, 2, values[1])
    sheet["A16"] = "Compatibility"
    sheet["B16"] = "Excel 2019 or Microsoft 365. Formulas recalculate when the workbook opens."
    sheet["A18"] = "Statistical testing"
    sheet["B18"] = (
        "No clinical hypothesis testing is claimed. The current tests verify "
        "calculation, temporal, boundary, missingness, reproducibility, and "
        "configuration behavior only."
    )
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 95
    for row in range(8, 20):
        sheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")


def style_headers(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical=cell.alignment.vertical or "top",
                        horizontal=cell.alignment.horizontal,
                    )
        for row_number in (7, 21, 23):
            if row_number <= sheet.max_row:
                for cell in sheet[row_number]:
                    if cell.value is not None:
                        cell.font = Font(bold=True, color=WHITE)
                        cell.fill = PatternFill("solid", fgColor=TEAL)


def main() -> None:
    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    patients = json.loads(DATA_PATH.read_text(encoding="utf-8"))
    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    build_readme(workbook)
    build_input(workbook, config)
    build_results(workbook)
    build_assumptions(workbook, config)
    build_cohort(workbook, patients)
    build_dictionary(workbook)
    build_review(workbook)
    style_headers(workbook)
    workbook.active = 0
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(f"Built {OUTPUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
